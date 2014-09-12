import argparse
import csv
import datetime
import sys

try:
    from cStringIO import StringIO
except ImportError:
    from StringIO import StringIO

from openpyxl import Workbook
from openpyxl.cell.cell import column_index_from_string
from openpyxl.styles import Style, Font, numbers
from openpyxl.writer.dump_worksheet import DumpWorksheet
from openpyxl.xml.functions import XMLGenerator, start_tag, end_tag, tag


class CSV2XLSX(object):
    """
    Converts CSV from stdin into XLSX to stdout. Also optionally converts columns to integers
    or datetimes and adjusts column widths.
    """
    def __init__(self):
        self.args = self.parse_args()

        self.integer_cols = [column_index_from_string(c)-1 for c in self.args.integer_cols.split(',')]\
            if self.args.integer_cols else []

        def datetime_spec_parse(arg_tuple):
            str_col, in_format, out_format = arg_tuple.split(';', 2)
            out_cell = {'style': Style(number_format=numbers.NumberFormat(format_code=out_format))}
            return column_index_from_string(str_col)-1, (in_format, out_cell)

        self.datetime_spec_dict = dict(datetime_spec_parse(a) for a in self.args.datetime_spec or ())

        self.HEADER_STYLE = Style(font=Font(bold=True))
        self.strptime = datetime.datetime.strptime

    def convert(self, infile, outfile):
        WidthsDumpWorksheet.col_widths = {}

        for width_spec in self.args.col_width or ():
            col, width = width_spec.split(',', 1)
            if '-' in col:
                col_lo, col_hi = map(column_index_from_string, col.split('-', 1))
            else:
                col_lo = col_hi = column_index_from_string(col)

            WidthsDumpWorksheet.col_widths[(col_lo, col_hi)] = width

        # TODO: This will be used when we switch to openpyxl 2.1.
        #Workbook._optimized_worksheet_class = WidthsDumpWorksheet

        wb = Workbook(optimized_write=True, optimized_worksheet_class=WidthsDumpWorksheet)

        ws = wb.create_sheet(title=self.args.sheet_name)

        for row in self.iter_rows(self.get_reader(infile)):
            ws.append(row)

        wb.save(outfile)

        # Clean up the mess.
        WidthsDumpWorksheet.col_widths = {}

    def get_reader(self, infile):
        return csv.reader(infile, delimiter=self.args.delimiter, quotechar=self.args.quotechar)

    def iter_rows(self, reader):
        for unused in range(self.args.header_rows):
            yield [{'value': cell.decode(self.args.input_encoding), 'style': self.HEADER_STYLE}
                   for cell in reader.next()]

        for row in reader:
            row = [cell.decode(self.args.input_encoding) for cell in row]

            for col in self.integer_cols:
                if row[col]:
                    row[col] = int(row[col])

            for col, (input_format, out_cell) in self.datetime_spec_dict.items():
                if row[col]:
                    out_cell['value'] = self.strptime(row[col], input_format)
                    row[col] = out_cell

            yield row

    def parse_args(self):
        p = argparse.ArgumentParser(description=self.__doc__)
        p.add_argument("input_encoding", help="The character encoding of the input CSV.")
        p.add_argument("sheet_name", help="The name to put on the single sheet.")

        p.add_argument("--delimiter", "-d", default=';',
                       help="Cell delimiter in the input. Default: semicolon (;)")
        p.add_argument("--quotechar", "-q", default='"',
                       help="Quoting character in the input. Default: double quote (\")")
        p.add_argument("--header-rows", "-H", type=int, default=0, metavar="NUMBER",
                       help="The number of leading rows to just pass with bold style "
                            "without any type conversions. Default: 0")

        p.add_argument("--integer-cols", "-i", metavar='COLUMNS',
                       help="Comma-delimited list of columns whose values should be converted "
                            "to integer. Example: -iA,G,AA")
        p.add_argument("--datetime-spec", "-t", action='append', metavar="COLUMN;I-FORMAT;O-FORMAT",
                       help="Points to one of the columns that should be treated as date, time "
                            "or datetime value. Semicolon-delimited tuple of column, "
                            "input datetime format and output custom datetime format. "
                            "Example: -t\"C;%%d.%%m.%%Y %%H:%%M;dd.mm.yyyy hh:mm\"")

        p.add_argument("--col-width", "-w", action='append', metavar="COLUMN,WIDTH",
                       help="Specifies the width for given column or column range. "
                            "Example: -wA,20 -wO-R,30")
        return p.parse_args()


# TODO: This will be used when we switch to openpyxl 2.1.
# class WidthsDumpWorksheet(DumpWorksheet):
#     """
#     This is a hack to get the <cols>...</cols> element to optimized writer by force.
#     """
#     def write_header(self):
#         doc = super(WidthsDumpWorksheet, self).write_header()
#
#         start_tag(doc, 'cols')
#         for (col_lo, col_hi), width in self.col_widths.items():
#             tag(doc, 'col', dict(min=str(col_lo), max=str(col_hi), width=width, customWidth="1"))
#         end_tag(doc, 'cols')
#
#         return doc


class WidthsDumpWorksheet(DumpWorksheet):
    """
    This is a hack to get the <cols>...</cols> element to optimized writer by force.
    """
    def write_header(self):

        fobj = self.get_temporary_file(filename=self._fileobj_header_name)
        doc = XMLGenerator(fobj)

        start_tag(doc, 'worksheet',
                  {
                      'xmlns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                      'xmlns:r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'})
        start_tag(doc, 'sheetPr')
        tag(doc, 'outlinePr',
                 {'summaryBelow': '1',
                 'summaryRight': '1'})
        end_tag(doc, 'sheetPr')
        tag(doc, 'dimension', {'ref': 'A1:%s' % (self.get_dimensions())})
        start_tag(doc, 'sheetViews')
        start_tag(doc, 'sheetView', {'workbookViewId': '0'})
        tag(doc, 'selection', {'activeCell': 'A1',
                 'sqref': 'A1'})
        end_tag(doc, 'sheetView')
        end_tag(doc, 'sheetViews')
        tag(doc, 'sheetFormatPr', {'defaultRowHeight': '15'})

        # START csv2xlsx
        start_tag(doc, 'cols')
        for (col_lo, col_hi), width in self.col_widths.items():
            tag(doc, 'col', dict(min=str(col_lo), max=str(col_hi), width=width, customWidth="1"))
        end_tag(doc, 'cols')
        # END csv2xlsx

        start_tag(doc, 'sheetData')


def main():
    output = StringIO()

    CSV2XLSX().convert(sys.stdin, output)

    sys.stdout.write(output.getvalue())
